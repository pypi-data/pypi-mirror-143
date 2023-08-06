import sys
import argparse

import openpyxl

from .vcard import VCARD
from .read_file import analysis


def makeVcard(ifn, ofn=None, ftpye="excel", sheet=0):
    '''
    ifn: 输入文件名
    ofn: 输出文件名
    '''
    if not ofn:
        ofn = "{}.{}".format(ifn.split(".")[0], "vcf")
    else:
        ofn = "{}.{}".format(ofn, "vcf") if len(ofn.split(".")) == 1 else ofn
    data = []
    for item in analysis(ifn, ftpye, sheet):
        data.append(VCARD().set_data(**item).to_string())
        # print(VCARD().set_data(**item).to_string())
        # print(item)
    with open(ofn, "w", encoding="utf-8") as f:
        for d in data:
            f.write(d)
    print(ofn, "文件生成完成！")
    return 0


def arguments_parse():
    parser = argparse.ArgumentParser(prog="make_vcard",
                                     description="基本用法 python -m make_vcard inputFileName.xlsx [-o outputFileName]",
                                     epilog="注意：不同设备对 字段 的支持不一样，可能您按照这个规则并不会得到您想要的结果！",
                                     )
    parser.add_argument("--example", action="store_true", help="生成示例excel文件；")
    parser.add_argument("--sheet", type=str, default="0", help="指定读取哪张表；默认为第一张表")
    parser.add_argument("--filetype", type=str, default="excel", help="指定文件类型（默认为excel）只可以指定为 csv")
    parser.add_argument("-o", type=str,
                        default="", help="输出保存的文件名；不需要加后缀！")

    parser.add_argument("inputFileName", default="",
                        type=str, help="指定读取的文件")
    return parser.parse_args()


def example():
    excel = openpyxl.Workbook()
    sheet = excel.active
    sheet.title = "基础版"
    sheet["A1"] = "姓名"
    sheet["A2"] = "张三"
    sheet["B1"] = "电话"
    sheet["B2"] = "13000000000"
    sheet["C1"] = "备注"
    sheet["C2"] = "备注备注"
    sheet["D1"] = "邮箱"
    sheet["D2"] = "123@456.com"
    # ============================
    sheet2 = excel.create_sheet("进阶版")
    vcard = {
        "前缀": "test_",
        "后缀": "后缀",
        "姓名": "李四",
        "昵称": "lisi",
        "备注": "备注备注",
        "生日": "2000-10-01",
        "组织": "正义组织",
        "职位": "做个好人",
        "电话": "150000000",
        "电话1": "170000000",
        "电话2(工作)": "170000000",
        "电话3（住宅）": "180000000",
        "邮箱": "123@aa.cc",
        "邮箱1": "1234@aa.cc",
        "邮箱2": "12345@aa.cc",
        "网址": "http://test.com",
        "地址": "中国--------------------",
    }
    i = 1
    for k in vcard:
        a = chr(64+i)
        sheet2[str(a)+"1"] = k
        sheet2[str(a)+"2"] = vcard[k]
        i += 1

    sheet3 = excel.create_sheet("食用说明")
    sheet3["A1"] = "说明"
    sheet3["A2"] = "只有 姓名和电话 是必要的，其它是可以任意组合"
    sheet3["A3"] = "有 电话, 邮箱, 网址, 地址 是可以重复出现的；其规则是 名称+数字；例: 电话A"
    sheet3["A4"] = "可重复出现的数据可定义其类型"
    sheet3["A5"] = "电话的类型有：优先，住宅，工作，单元，传真，其它 "
    sheet3["A6"] = "邮箱的类型有：住宅，工作，其它"
    sheet3["A7"] = "网址的类型有：住宅，工作，其它 "
    sheet3["A8"] = "地址的类型有：住宅，工作，其它，邮寄 "
    sheet3["A9"] = "例如： 邮箱(住宅)"
    sheet3["A10"] = "例如： 网址（工作）"
    sheet3["A11"] = "注意：不同设备对 字段 的支持不一样，可能您按照这个规则并不会得到您想要的结果！"

    sheet4 = excel.create_sheet("必要的")
    sheet4["A1"] = "姓名"
    sheet4["A2"] = "张三"
    sheet4["B1"] = "电话"
    sheet4["B2"] = "13000000000"

    excel.save("example.xlsx")
    print("示例文件已成功生成。")
    return 0


def cmd():
    if len(sys.argv) == 2 and sys.argv[1] == "--example":
        sys.exit(example())

    args = arguments_parse()
    SHEET = 0
    # print(args.sheet, type(args.sheet))
    if args.sheet != "0":
        SHEET = args.sheet
        
    makeVcard(args.inputFileName, args.o, args.filetype, SHEET)
