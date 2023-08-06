from xltpl.writerx import BookWriter
from xltpl.writer import BookWriter as xlsBookWriter
from openpyxl import load_workbook
    

class xlsxDocument:
    def __init__(self, input_path, data=None, jinja_env=None):
        self.input_path =input_path
        self._bookWriter = BookWriter(input_path)
        self._init(data, jinja_env)

    def _init(self, data=None, jinja_env=None):
        if data:
            self.data = data
        if jinja_env:
            self._update_jinja_env(jinja_env)

    def render(self, data=None, jinja_env=None):
        self._init(data, jinja_env)
        
        for sheet in load_workbook(self.input_path).sheetnames:
            # ws: Worksheet = wb[wb.sheetnames[0]]
            self.render_sheet(sheet)

    def save(self, file):
        self._bookWriter.save(file)

    def render_sheet(self, sheet_name):
        sheet_resource = self._bookWriter.sheet_resource_map.get(sheet_name)
        sheet_writer = self._bookWriter.get_sheet_writer(sheet_resource, sheet_name)
        sheet_resource.render_sheet(sheet_writer, self.data)

    def _update_jinja_env(self, jinja_env):
        if jinja_env:
            for attr in ["extensions", "filters", "globals"]:
                setattr(
                    self._bookWriter.jinja_env,
                    attr,
                    getattr(self._bookWriter.jinja_env, attr)
                    | getattr(jinja_env, attr),
                )

class xlsDocument(xlsxDocument):
     def __init__(self, input_path, data=None, jinja_env=None):
        xlsxDocument.__init__(self,input_path, data=None, jinja_env=None)
        self._bookWriter =xlsBookWriter(input_path)
