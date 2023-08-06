import logging
import pandas as pd
import csv
import xlrd
from contextlib import closing
from openpyxl import load_workbook
from airflow.models import BaseOperator
from airflow.utils.decorators import apply_defaults
from airflow.exceptions import AirflowException
from airflow.hooks.base_hook import BaseHook

class NovigiExcelToCSVExportOperator(BaseOperator):

    template_fields = ['input_excel']

    def __init__(self, input_excel: str, output_csv: str, sheet_name: str = None, column_range: range = None, row_range: range = None, headers: list = None, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self.input_excel = input_excel
            self.output_csv = output_csv
            self.sheet_name = sheet_name
            self.column_range = column_range
            self.row_range = row_range
            self.headers = headers

    def clean(self, value):
            #use this callback to clean
            return value

    def execute(self, context):

            format_01 = ".xlsx"
            format_02 = ".xls"

            if format_01 in self.input_excel:
                wb = load_workbook(filename = self.input_excel)
                if self.sheet_name is None:
                    sheet = wb.active
                else:
                    sheet = wb[self.sheet_name]

                if self.column_range is None:
                    self.column_range = range(1, sheet.max_column + 1)

                if self.row_range is None:
                    self.row_range = range(1, sheet.max_row + 1)

                with open(self.output_csv, "w") as outfile:
                    writer = csv.writer(outfile, quoting=csv.QUOTE_NONNUMERIC)
                    if self.headers is not None:
                        writer.writerow(headers)
                    for i in self.row_range:
                        row = []
                        for j in self.column_range:
                            cell = sheet.cell(row=i, column=j)
                            row.append(cell.value)

                        writer.writerow([self.clean(x) for x in row])
                    outfile.close()

            elif format_02 in self.input_excel:

                data_xls = pd.read_excel(self.input_excel, self.sheet_name, dtype=str, index_col=None)
                data_xls.to_csv(self.output_csv, encoding='utf-8', index=False)

            else:
                raise ValueError('Particular Excel file type is not supprted when it is converting to a csv file! Supported formats are .xls and .xlsx')

            return True
