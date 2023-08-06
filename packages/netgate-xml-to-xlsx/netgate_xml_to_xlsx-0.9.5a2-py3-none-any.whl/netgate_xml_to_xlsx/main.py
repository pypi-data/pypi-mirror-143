"""Main netgate converstion module."""
# Copyright © 2022 Appropriate Solutions, Inc. All rights reserved.

import argparse
from collections import OrderedDict
import datetime
import html
import ipaddress
import os
from pathlib import Path
import re
import sys

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter
import xmltodict

from ._version import __version__


class ScriptError(Exception):
    """Generic script error."""


class UnknownField(ScriptError):
    """Parsing something new."""


class MissingField(ScriptError):
    """Expected field not found."""


def parse_args():
    """
    Parse command line arguments.

    Process in_files and out_dir.
    """
    parser = argparse.ArgumentParser("Netgate XML to XLSX")

    default = "./output"
    parser.add_argument(
        "--output-dir",
        "-o",
        type=str,
        default=default,
        help=f"Output directory. Default: {default}",
    )
    parser.add_argument(
        "in_files", nargs="+", help="One or more Netgate .xml files to process."
    )
    parser.add_argument(
        "--sanitize",
        action="store_true",
        help="Sanitize the input xml files and save as <filename>-sanitized.",
    )
    parser.add_argument(
        "--version",
        action="version",
        version=f"{__version__}",
        help="Show version number.",
    )

    args = parser.parse_args()

    # Filter files in/out.
    if args.sanitize:
        args.in_files = _filter_infiles(args.in_files, include=False)
        msg = "All files already sanitized."
    else:
        args.in_files = _filter_infiles(args.in_files)
        msg = (
            "No files contain 'sanitized' in the name.\n"
            "Run --sanitize before processing."
        )

    if not args.in_files:
        print(msg)
        exit(-1)

    # Convert list of in_files to list of Path objects.
    # Ensure they are files, not directories.
    args.in_files = [Path(x) for x in args.in_files]
    args.in_files = [x for x in args.in_files if x.is_file()]

    # Convert output-dir to path and optionally create path.
    out_dir = Path(args.output_dir)
    try:
        out_dir.mkdir(parents=True, exist_ok=True)
        args.output_dir = out_dir
    except OSError as err:
        print(f"Error: {err}")
        sys.exit(-1)
    return args


def _sanitize_xml(raw_xml: str) -> str:
    """Sanitize the xml."""
    regexes = (
        re.compile("(<bcrypt-hash>).*?(</bcrypt-hash>)"),
        re.compile("(<radius_secret>).*?(</radios_secret>)"),
        re.compile("(<lighttpd_ls_password>).*?(</lighttpd_ls_password>)"),
        re.compile("(<stats_password>).*?(</stats_password>)"),
        re.compile("(<password>).*?(</password>)"),
        re.compile("(<tls>).*?(</tls>)"),
    )
    for regex in regexes:
        raw_xml = regex.sub(r"\1SANITIZED\2", raw_xml)
    return raw_xml


def _format_privs(privs: list[tuple[str, str | None]] | str | None) -> str | None:
    """Format privileges into a string paragraph.

    Single privileges are presented as a string.
    Multiple privileges as a list of string.
    """
    if privs is None:
        return None
    assert privs is not None

    if isinstance(privs, str):
        return privs

    privs.sort(key=lambda x: x.casefold())
    return "\n".join(privs)


def _unescape(value: str | None) -> str | None:
    """Unescape XML entities."""
    if value is None:
        return value
    assert value is not None

    return html.unescape(value)


def _adjust_field_value(*, field_name: str, value: str | None) -> str | None:
    """Make adjustments based on field_name."""
    if value is None:
        return None
    assert value is not None

    # Convert XML escape codes
    assert value is not None

    if field_name == "descr":
        value = value.replace("<br />", "\n")
        lines = [x.strip() for x in value.split("\n")]
        return "\n".join(lines)

    # May be specific only to our environment. Details divided by ||
    if field_name == "detail":
        value = value.replace("||", "\n")
        lines = [x.strip() for x in value.split("\n")]
        return "\n".join(lines)

    if (
        field_name
        in "data_ciphers,local_network,local_networkv6,remote_network,remote_networkv6".split(
            ","
        )
    ):
        values = [x.strip() for x in value.split(",")]
        return "\n".join(values)

    if field_name in "custom_options".split(","):
        values = [x.strip() for x in value.split(";")]
        return "\n".join(values)

    if field_name in "shared_key".split(",") and len(value) > 30:
        # Sanitize the value
        return f"{value[:20]}\n......\n{value[-20:]}"

    return value


def _updated_or_created(node: OrderedDict) -> str:
    """Return "updated" or "created" value, or ""."""
    if updated := _get_element(node, "updated,time"):
        return updated
    return _get_element(node, "created,time")


def _get_element(
    root_node: OrderedDict, els: list[str] | str, default=""
) -> OrderedDict | str | None:
    """
    Iterate down the tree and return path.

    Use try/except for missing keys as None is a valid return value.
    """
    if isinstance(els, str):
        els = els.split(",")

    if root_node is None:
        return default

    node = root_node
    try:
        for el in els:
            node = node[el]
            if node is None:
                return default
        return node
    except KeyError:
        return default


def _load_standard_nodes(
    *, nodes: OrderedDict | list | None, field_names: list[str]
) -> list[list]:
    """Load nodes that do not require special handling into rows."""
    rows = []
    if nodes is None:
        return rows
    assert nodes is not None

    # If a single dictionary, put it into a list
    if isinstance(nodes, OrderedDict):
        nodes = [nodes]

    for node in nodes:
        if node is None:
            # blank <openvpn-server></openvpn-server> for example.
            continue
        row = []
        for field_name in field_names:
            row.append(
                _adjust_field_value(
                    field_name=field_name, value=node.get(field_name, "")
                )
            )
        rows.append(row)
    return rows


class PfSense:
    """Handle all pfSense parsing and conversion."""

    def __init__(self, args: argparse.Namespace, in_filename: str) -> None:
        """
        Initialize and load XML.

        Technically a bit too much work to do in an init (since it can fail).
        """
        self.args = args
        self.in_file = Path(in_filename)
        self.raw_xml: dict = {}
        self.pfsense: dict = {}
        self.workbook: Workbook = Workbook()

        # ss_filename is expected to be overwritten by
        self.ss_filename = "output.xlxs"
        self._init_styles()
        self.default_alignment = Alignment(wrap_text=True, vertical="top")

        self._load()

    def _init_styles(self) -> None:
        """Iniitalized worksheet styles."""
        xlsx_header_font = Font(name="Calibri", size=16, italic=True, bold=True)
        xlsx_body_font = Font(name="Calibri", size=16)
        xlsx_footer_font = Font(name="Calibri", size=12, italic=True)

        body_border = Border(
            bottom=Side(border_style="dotted", color="00000000"),
            top=Side(border_style="dotted", color="00000000"),
            left=Side(border_style="dotted", color="00000000"),
            right=Side(border_style="dotted", color="00000000"),
        )

        alignment = Alignment(wrap_text=True, vertical="top")

        header = NamedStyle(name="header")
        header.alignment = alignment
        header.fill = PatternFill(
            "lightTrellis", fgColor="00339966"
        )  # fgColor="000000FF")  #fgColor="0000FF00")
        header.font = xlsx_header_font
        header.border = Border(
            bottom=Side(border_style="thin", color="00000000"),
            top=Side(border_style="thin", color="00000000"),
            left=Side(border_style="dotted", color="00000000"),
            right=Side(border_style="dotted", color="00000000"),
        )

        normal = NamedStyle(name="normal")
        normal.alignment = alignment
        normal.border = body_border
        normal.fill = PatternFill("solid", fgColor="FFFFFFFF")
        normal.font = xlsx_body_font

        footer = NamedStyle("footer")
        footer.alignment = Alignment(wrap_text=False, vertical="top")
        footer.border = body_border
        normal.fill = PatternFill("solid", fgColor="FFFFFFFF")
        footer.font = xlsx_footer_font

        self.workbook.add_named_style(header)
        self.workbook.add_named_style(normal)
        self.workbook.add_named_style(footer)

    def _load(self) -> OrderedDict:
        """Load and parse Netgate xml firewall configuration.

        Return pfsense keys.
        """
        # with open(self.in_file, encoding="utf-8") as fh:
        self.raw_xml = self.in_file.read_text(encoding="utf-8")
        # self.raw_xml = fh.read()
        data = xmltodict.parse(self.raw_xml)
        self.pfsense = data["pfsense"]

    def _write_sheet(
        self,
        *,
        name: str,
        field_names: list[str],
        column_widths: list[int],
        rows: list[list],
    ):
        sheet = self.workbook.create_sheet(name)
        self._sheet_header(sheet, field_names, column_widths)

        # Define starting row num in case there are no rows to display.
        row_num = 2
        for row_num, row in enumerate(rows, start=row_num):
            self._write_row(sheet, row, row_num)
        self._sheet_footer(sheet, row_num)

    def _write_row(
        self, sheet: Worksheet, row: list, row_num: int, style_name: str = "normal"
    ) -> None:
        """
        Write a row into the spreadsheet.

        Args:
            row: A list of values to write into the row.

            row_increment: Number of rows to increment in spreadsheet before writing.

            style_name: Named XLSX style.

        Always increment the row before writing.

        """
        for column_number, value in enumerate(row, start=1):
            column_letter = get_column_letter(column_number)
            coordinate = f"{column_letter}{row_num}"
            sheet[coordinate] = value
            sheet[coordinate].style = style_name

    def _sheet_header(self, sheet: Worksheet, columns: list, column_widths: list[int]):
        """Write header row then set the column widths."""
        self._write_row(sheet, columns, 1, "header")

        for column_number, width in enumerate(column_widths, start=1):
            column_letter = get_column_letter(column_number)
            sheet.column_dimensions[column_letter].width = width

    def _sheet_footer(self, sheet: Worksheet, row_number: int) -> None:
        """Write footer information on each sheet."""
        now = datetime.datetime.now().strftime("%Y-%m-%dT%H:%M")
        run_date = f"Run date: {now}"

        self._write_row(sheet, [run_date], row_number + 1, style_name="footer")

    def sanitize(self) -> None:
        """
        Sanitize the raw XML and save as original filename + '-sanitized'.

        The Netgate configuration file XML is well ordered and thus searchable via regex.
        """
        self.raw_xml = _sanitize_xml(self.raw_xml)

        # Save sanitized XML
        parts = os.path.splitext(self.in_file)
        if len(parts) == 1:
            out_path = Path(f"{parts[0]}-sanitized")
        else:
            out_path = Path(f"{parts[0]}-sanitized{parts[1]}")
        out_path.write_text(self.raw_xml, encoding="utf-8")
        print(f"Sanitized file written: {out_path}.")

        # Delete the unsanitized file.
        self.in_file.unlink()
        print(f"Deleted original file: {self.in_file}.")

    def save(self) -> None:
        """Delete empty first sheet and then save Workbook."""
        sheets = self.workbook.sheetnames
        del self.workbook[sheets[0]]
        out_path = self.args.output_dir / self.ss_filename
        self.workbook.save(out_path)

    def _nice_address_sort(self, data: str) -> str:
        """Sort addresses that may consist of domains and IPv4/v6 addresses."""
        addresses = [x.strip() for x in data.split(" ")]
        numeric = [x for x in addresses if len(x) > 1 and x[0] in ("0123456789")]
        numeric.sort(key=lambda x: ipaddress.ip_address(x.split("/")[0]))

        non_numeric = [
            x for x in addresses if len(x) > 1 and x[0] not in ("0123456789")
        ]
        non_numeric.sort(key=str.casefold)

        result = []
        result.extend(non_numeric)
        result.extend(numeric)
        return "\n".join(result)

    def _clean_split(self, data: str, split_char: str = ",") -> str:
        """Convert str to list of strings.

        Split string on split_char.
        Trim whitespace.
        Remove blank lines.
        """
        lines = [x.strip() for x in data.split(split_char)]
        lines = [x for x in lines if x]
        return "\n".join(lines)

    def system(self) -> None:
        """
        Sheet with system-level information.

        Only showing interesting information (at least to me at the moment).
        """
        rows = []
        field_names = "name,value".split(",")
        column_widths = [int(x) for x in "40,40".split(",")]

        node = self.pfsense
        for key in "version,lastchange".split(","):
            rows.append([key, _unescape(node.get(key, ""))])

        # Check version number.
        if (version := int(float(rows[0][1]))) != 21:
            assert version is not None
            print(
                f"Warning: File uses version {version}.x. "
                "Script is only tested on version 21 XML formats."
            )

        node = self.pfsense["system"]
        for key in "optimization,hostname,domain,timezone".split(","):
            rows.append([key, _unescape(node.get(key, ""))])

        # Ugly getting this twice.
        self.ss_filename = (
            f"""{node.get("hostname", "")}.{node.get("domain", "")}.xlsx"""
        )

        time_servers = "\n".join(node.get("timeservers", "").split(" "))
        rows.append(["timeservers", time_servers])

        rows.append(["bogons", _get_element(node, "bogons,interval", "TBD")])
        rows.append(["ssh", _get_element(node, "ssh,enabled", "TBD")])
        rows.append(["dnsserver", "\n".join(node.get("dnsserver", ""))])

        self._write_sheet(
            name="System",
            field_names=field_names,
            column_widths=column_widths,
            rows=rows,
        )

    def system_groups(self) -> None:
        """
        Sheet with system.group information.

        Multiple groups with multiple privileges.
        Display privileges alpha sorted.
        """
        rows = []
        field_names = "name,description,scope,gid,priv".split(",")
        column_widths = [int(x) for x in "40,80,20,20,80".split(",")]

        nodes = _get_element(self.pfsense, "system,group")
        if not nodes:
            return
        if isinstance(nodes, OrderedDict):
            # Only found one.
            nodes = [nodes]

        nodes.sort(key=lambda x: x["name"].casefold())

        for node in nodes:
            row = []
            for key in field_names:
                value = _unescape(node.get(key, ""))
                if key == "priv":
                    value = _format_privs(value)
                row.append(value)
            rows.append(row)

        self._write_sheet(
            name="System Groups",
            field_names=field_names,
            column_widths=column_widths,
            rows=rows,
        )

    def system_users(self) -> None:
        """
        Sheet with system.user information.

        Not all fields displayed as # column on dashboard and webguicss are uninteresting
        (at least to me at the moment).
        """
        rows = []
        field_names = "name,descr,scope,expires,ipsecpk,uid,cert".split(",")
        column_widths = [int(x) for x in "40,60,20,20,20,10,60".split(",")]

        nodes = _get_element(self.pfsense, "system,user")
        if not nodes:
            return
        if isinstance(nodes, OrderedDict):
            # Only found one.
            nodes = [nodes]
        nodes.sort(key=lambda x: x["name"].casefold())

        for node in nodes:
            row = []
            for key in field_names:
                row.append(_unescape(node.get(key, "")))
            rows.append(row)

        self._write_sheet(
            name="System Users",
            field_names=field_names,
            column_widths=column_widths,
            rows=rows,
        )

    def aliases(self) -> None:
        """Create the aliases sheet."""
        rows = []
        field_names = "name,type,address,url,updatefreq,descr,detail".split(",")
        column_widths = [int(x) for x in "40,40,40,80,20,80,80".split(",")]

        nodes = _get_element(self.pfsense, "aliases,alias")
        if not nodes:
            return
        if isinstance(nodes, OrderedDict):
            # Only found one.
            nodes = [nodes]
        nodes.sort(key=lambda x: x["name"].casefold())

        rows = _load_standard_nodes(nodes=nodes, field_names=field_names)
        self._write_sheet(
            name="Aliases",
            field_names=field_names,
            column_widths=column_widths,
            rows=rows,
        )

    def _updated_or_created(self, data: OrderedDict) -> str:
        """Return updated or created timestamp."""
        if updated := _get_element(data, "updated,time"):
            return updated
        return _get_element(data, "created,time")

    def _rules_source(self, row: list, field_index: int) -> str:
        """Extract source or 'any'."""
        source = row[field_index]
        if "any" in source:
            return "any"

        if "address" in source:
            return source["address"]

        if "network" in source:
            return source["network"]

        raise UnknownField(f"Unknown filter/rules/source field: {source.keys()}")

    def _rules_destination(self, row: list, field_index: int) -> str:
        """Extract destination address and port."""
        destination = row[field_index]
        if "any" in destination:
            return "any"

        if "network" in destination:
            address = destination["network"]
        elif "address" in destination:
            address = destination["address"]
        else:
            raise UnknownField(
                f"Destination missing address or network: {destination.keys()}"
            )

        try:
            port = destination["port"]
            return f"{address}:{port}"
        except KeyError:
            # Aliases don't always have ports.
            return address

        return f"{address}:{port}"

    def _rules_username_time(self, row: list, field_index: int) -> str:
        """Extract username/time from created or updated."""
        field = row[field_index]

        # Not all records have an 'updated' so return "" if field is missing.
        if not field:
            return ""
        username = field.get("username")
        username = username if username is not None else ""

        time = field.get("time")
        if time is not None:
            timestamp = datetime.datetime.fromtimestamp(int(time)).strftime(
                "%Y-%m-%d %H-%M-%S"
            )
        else:
            timestamp = ""
        return f"{username} {timestamp}"

    def rules(self) -> None:
        """Create the rules sheet."""
        rows = []
        field_names = (
            "id,tracker,type,interface,ipprotocol,tag,tagged,max,max_src_nodes,"
            "max_src-conn,max-src-states,statetimeout,statetype,os,source,destination,"
            "log,descr,created,updated"
        ).split(",")
        column_widths = [
            int(x)
            for x in (
                "10,20,10,15,15,10,10,10,20,20,20,20,30,50,50,50,10,80,85,85".split(",")
            )
        ]
        source_index = field_names.index("source")
        destination_index = field_names.index("destination")
        created_index = field_names.index("created")
        updated_index = field_names.index("updated")

        nodes = _get_element(self.pfsense, "filter,rule")
        if not nodes:
            return
        if isinstance(nodes, OrderedDict):
            # Only found one.
            nodes = [nodes]
        # Sort rules so that latest changes are at the top.
        nodes.sort(
            key=_updated_or_created,
            reverse=True,
        )

        for node in nodes:
            row = []
            for field_name in field_names:
                row.append(_unescape(node.get(field_name, "")))
            row = ["" if x is None else x for x in row]

            row[source_index] = self._rules_source(row, source_index)
            row[destination_index] = self._rules_destination(row, destination_index)
            row[created_index] = self._rules_username_time(row, created_index)
            row[updated_index] = self._rules_username_time(row, updated_index)

            rows.append(row)

        self._write_sheet(
            name="Rules",
            field_names=field_names,
            column_widths=column_widths,
            rows=rows,
        )

    def interfaces(self) -> None:
        """
        Document all interfaces.

        TODO: Review blockbogons. Does existence == On?
        """
        rows = []
        # Prepend 'name' before calling _write_sheet.
        field_names = (
            "descr,alias-address,alias-subnet,spoofmac,enable,"
            "blockpriv,blockbogons,ipaddr,subnet,gateway"
        ).split(",")
        column_widths = [int(x) for x in "20,40,20,20,20,20,20,20,20,40".split(",")]

        # Don't sort interfaces. Want them in the order they are encountered.
        # Interfaces is an OrderedDict
        nodes = _get_element(self.pfsense, "interfaces")
        if not nodes:
            return
        # In this case we want a single OrderedDict.
        # Remove 'name' from the field_names as we're going to replace that with the key.
        del field_names[0]

        for name, node in nodes.items():
            row = [name]
            for field_name in field_names:
                row.append(_unescape(node.get(field_name, "")))
            rows.append(row)
        field_names.insert(0, "name")

        self._write_sheet(
            name="Interfaces",
            field_names=field_names,
            column_widths=column_widths,
            rows=rows,
        )

    def gateways(self) -> None:
        """Document gateways."""
        rows = []
        # Append 'defaultgw4' and 'defaultgw6' before displaying.
        field_names = (
            "name,descr,interface,gateway,weight,ipprotocol,monitor_disable"
        ).split(",")
        column_widths = [int(x) for x in "20,40,20,20,10,30,30,20,20".split(",")]

        # Load default IPV4 and IPV6 gateways.
        # Don't want "None" for default gateway.
        default_gw4 = _get_element(self.pfsense, "gateways,defaultgw4", "")
        default_gw6 = _get_element(self.pfsense, "gateways,defaultgw6", "")

        # Which column has the gateway name.
        gw_name_col = 0

        # Don't sort nodes for now. Leave in order found.
        nodes = _get_element(self.pfsense, "gateways,gateway_item")
        if not nodes:
            return
        if isinstance(nodes, OrderedDict):
            # Only found one.
            nodes = [nodes]

        for node in nodes:
            row = []
            for field_name in field_names:
                try:
                    row.append(_unescape(node.get(field_name, "")))
                except AttributeError as err:
                    print(err)
                    sys.exit(-1)
            if default_gw4 == row[gw_name_col]:
                row.append("YES")
            else:
                row.append(None)
            if default_gw6 == row[gw_name_col]:
                row.append("YES")
            else:
                row.append(None)
            rows.append(row)

        field_names.extend(["defaultgw4", "defaultgw6"])
        self._write_sheet(
            name="Gateways",
            field_names=field_names,
            column_widths=column_widths,
            rows=rows,
        )

    def openvpn_server(self) -> None:
        """Document all OpenVPN servers."""
        rows = []
        field_names = (
            "vpnid,disable,mode,protocol,dev_mode,interface,ipaddr,local_port,"
            "description,custom_options,shared_key,digest,engine,tunnel_network,"
            "tunnel_networkv6,remote_network,remote_networkv6,gwredir,gwredir6,"
            "local_network,local_networkv6,maxclients,compression,compression_push,passtos,"
            "client2client,dynamic_ip,topology,serverbridge_dhcp,serverbridge_interface,"
            "serverbridge_routegateway,serverbridge_dhcp_start,serverbridge_dhcp_end,"
            "username_as_common_name,exit_notify,sndrcvbuf,netbios_enable,netbios_ntype,"
            "netbios_scope,create_gw,verbosity_level,ncp_enable,ping_method,keepalive_interval,"
            "keepalive_timeout,ping_seconds,ping_push,ping_action,ping_action_seconds,"
            "ping_action_push,inactive_seconds,data_ciphers,data_ciphers_fallback"
        ).split(",")
        column_widths = [
            int(x)
            for x in (
                "20,20,20,20,20,30,20,20,30,20,"  # 10
                "40,20,20,30,30,30,30,20,20,30,"  # 20
                "20,20,20,30,20,20,20,20,40,40,"  # 30
                "40,40,40,50,20,20,20,20,20,20,"  # 40
                "20,20,20,30,30,20,20,20,30,30,"  # 50
                "20,20,50"
            ).split(",")
        ]

        # Don't sort OpenVPN Servers. Want them in the order they are encountered.
        # Interfaces is an OrderedDict
        nodes = _get_element(self.pfsense, "openvpn,openvpn-server")
        if not nodes:
            return
        if isinstance(nodes, OrderedDict):
            nodes = [nodes]

        rows = _load_standard_nodes(nodes=nodes, field_names=field_names)
        self._write_sheet(
            name="OpenVPN Server",
            field_names=field_names,
            column_widths=column_widths,
            rows=rows,
        )

    def installed_packages(self) -> None:
        """Document all installed packages. Sort by name."""
        rows = []
        field_names = (
            "name,internal_name,descr,version,configuration_file,include_file,"
            "website,pkginfolink,filter_rule_function"
        ).split(",")
        column_widths = [int(x) for x in "40,40,50,20,50,50,80,80,50".split(",")]

        nodes = _get_element(self.pfsense, "installedpackages,package")
        if not nodes:
            return
        if isinstance(nodes, OrderedDict):
            # Only found one.
            nodes = [nodes]
        nodes.sort(key=lambda x: x["name"].casefold())

        rows = _load_standard_nodes(nodes=nodes, field_names=field_names)
        self._write_sheet(
            name="Installed Packages",
            field_names=field_names,
            column_widths=column_widths,
            rows=rows,
        )

    def unbound(self) -> None:
        """Document unbound elements."""
        rows = []

        # field names for acquiring information
        field_names = (
            "enable,active_interface,outgoing_interface,custom_options,custom_options,"
            "hideversion,dnssecstripped,port,system_domain_local_zone_type,sslcertref,"
            "dnssec,tlsport"
        ).split(",")
        column_widths = [
            int(x) for x in "20,20,20,20,20,20,20,20,20,20,20,20,80,80".split(",")
        ]
        node = _get_element(self.pfsense, "unbound")
        rows = _load_standard_nodes(nodes=node, field_names=field_names)

        # Only expect one row returned.
        assert len(rows) <= 1

        if not rows:
            # No unbound values. Nothing to output.
            return

        # Load multi-element items.
        domain_overrides_fieldnames = "domain,ip,descr,tls_hostname".split(",")
        domain_overrides = _load_standard_nodes(
            nodes=_get_element(node, "domainoverrides"),
            field_names=domain_overrides_fieldnames,
        )

        hosts_fieldnames = "host,domain,ip,descr,aliases".split(",")
        hosts = _load_standard_nodes(
            nodes=_get_element(node, "hosts"), field_names=hosts_fieldnames
        )

        subrows = []
        for domain_override in domain_overrides:
            zipped = OrderedDict(zip(domain_overrides_fieldnames, domain_override))
            subrows.append(
                "\n".join([f"{key}: {value}" for key, value in zipped.items()])
            )

        rows[0].append("\n\n".join(subrows))

        subrows = []
        for host in hosts:
            zipped = OrderedDict(zip(hosts_fieldnames, host))
            subrows.append(
                "\n".join([f"{key}: {value}" for key, value in zipped.items()])
            )

        rows[0].append("\n\n".join(subrows))

        # Add the two subrows columns to the field names.
        field_names.extend(("domainoverrides", "hosts"))

        self._write_sheet(
            name="Unbound",
            field_names=field_names,
            column_widths=column_widths,
            rows=rows,
        )


def banner(pfsense: PfSense) -> None:
    """Tell people what we're doing."""
    print(f"Output will be: {pfsense.args.output_dir / pfsense.ss_filename}.")


def _filter_infiles(in_files: list[str], include=True) -> list[str]:
    """
    Return list of in_files that contain (include) or do not contain (exclused) 'sanitized'.

    """
    if include:
        return [x for x in in_files if "sanitized" in x]
    return [x for x in in_files if "sanitized" not in x]


def main() -> None:
    """Driver."""
    args = parse_args()

    in_files = args.in_files

    for in_filename in in_files:
        pfsense = PfSense(args, in_filename)

        if args.sanitize:
            pfsense.sanitize()
            continue

        # Worksheet creation order.
        pfsense.system()

        # Need to parse system before banner has the information it needs.
        banner(pfsense)
        pfsense.system_groups()
        pfsense.system_users()
        pfsense.aliases()
        pfsense.rules()
        pfsense.interfaces()
        pfsense.gateways()
        pfsense.openvpn_server()
        pfsense.installed_packages()
        pfsense.unbound()
        pfsense.save()


if __name__ == "__main__":
    main()
