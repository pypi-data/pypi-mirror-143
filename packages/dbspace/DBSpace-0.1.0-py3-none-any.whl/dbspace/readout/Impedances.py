#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Feb 19 16:39:29 2018

@author: virati
"""

import sys
#sys.path.append('/home/virati/Dropbox/projects/Research/MDD-DBS/Ephys/IntegratedAnalysis/')
sys.path.append('/home/virati/Dropbox/projects/Research/MDD-DBS/Ephys/DBSpace/')
#import DBSpace as dbs
import DBSpace as dbo
from DBSpace import nestdict

import ipdb

import itertools as itt

import numpy as np
import scipy.signal as sig
import matplotlib.pyplot as plt
import pandas as pd

import seaborn as sns

sns.set_context('paper')
sns.set(font_scale=2)
sns.set_style('white')

from openpyxl import load_workbook

# Create a function to calculat the total brain volume
def total_brain_volume(gm, wm, wsf):
    # returns the sum of gm, wm, wsf
    return gm + wm + wsf

# Create a function to calculat the total brain volume
def gm_total_ratio(gm, total):
    # returns the ratio of gm to total brain volume
    return gm / total        
        
class Anatomy:
    def __init__(self):
        lead = nestdict()
        
        lead['Left'] = pd.read_csv('/home/virati/Dropbox/projects/Research/MDD-DBS/Data/Anatomy/CT/VTRVTA_6AMP_1.csv',  names=['Subj', 'Contact', 'GM','WM', 'CSF'])
        lead['Right'] = pd.read_csv('/home/virati/Dropbox/projects/Research/MDD-DBS/Data/Anatomy/CT/VTRVTA_6AMP_2.csv',  names=['Subj', 'Contact', 'GM', 'WM','CSF'])
        
        for side in ['Left','Right']:
            # Create a total_volume variable that is the output of the function
            lead[side]['Total'] = total_brain_volume(lead[side]['GM'], lead[side]['WM'], lead[side]['CSF'])
            lead[side]['Gm_Total_Ratio'] = gm_total_ratio(lead[side]['GM'], lead[side]['Total'])


class Z_class:
    def __init__(self):
        #fname = '/home/virati/Downloads/MDT_Zs.xlsx'
        fname = '/home/virati/Dropbox/projects/Research/MDD-DBS/Data/Anatomy/CT/Zs_GMs_Es.xlsx'
        wb = load_workbook(fname)
        ws = wb['Zs']
        self.xs = nestdict()
        self.pts = ['DBS901','DBS903','DBS905','DBS906','DBS907','DBS908']
        for pp,pt in enumerate(self.pts):
            first_col = (pp)*5
            print(first_col)
            self.xs[pt]['Date'] = np.array([r[first_col].value for r in ws.iter_rows()])[1:]
            self.xs[pt]['Left'] = np.array([r[first_col+2].value for r in ws.iter_rows()])[1:]
            self.xs[pt]['Right'] = np.array([r[first_col+4].value for r in ws.iter_rows()])[1:]
            
            self.xs[pt]['Date'] = self.xs[pt]['Date'][self.xs[pt]['Date'] != None]
            
            #F ORDER IS CORRECT, gives us 4xobs matrix for each electrode
            self.xs[pt]['Left'] = self.xs[pt]['Left'].reshape(4,-1,order='F')
            self.xs[pt]['Right'] = self.xs[pt]['Right'].reshape(4,-1,order='F')
            #self.xs[pt]['Left'] = self.xs[pt]['Left'][self.xs[pt]['Left'] != None].reshape(4,-1,order='F')
            #self.xs[pt]['Right'] = self.xs[pt]['Right'][self.xs[pt]['Right'] != None].reshape(4,-1,order='F')
            
        self.get_recZs()
        self.gen_Zdiff()
            
    def ret_Z(self,pt,etrode):
        if etrode > 4:
            side = 'Right'
            etrode = etrode - 8
        else:
            side = 'Left'
            
        return self.xs['DBS'+pt][side][etrode::4]
    
    def load_pt_OnT_stim(self):
        pt_OnT = [(2,1),(2,2),(2,1),(2,2),(1,1),(2,1)]
        #self.pt_OnT = {pt:pt_OnT[pp] for pp,pt in enumerate(self.do_pts)}
        return pt_OnT
    
    def get_recZs(self):
        on_t_es = self.load_pt_OnT_stim()
        
        rec_Zs = {'Left':np.zeros((6,2,28)),'Right':np.zeros((6,2,28))}
                
        for pp, pt in enumerate(self.pts):
            print(pt)
            for side in ['Left','Right']:
            #RIGHT NOW THIS DOES E1 FIRST then E3; so goes from lower number to higher number
            #NEEDS TO BE PLOTTED BACKWARDS
                rec_Zs[side][pp,:,:] = np.vstack((self.xs[pt][side][on_t_es[pp][0]-1][0:28],self.xs[pt]['Left'][on_t_es[pp][0]+1][0:28]))
        
        
                rec_Zs[side][rec_Zs[side] > 4000] = np.nan
        
        #self.rec_Zs = rec_Zs
        return rec_Zs
            
    def load_pt_OnT_rec(self):
        pt_recEs = np.array((6,2,2))
        pt_recOnTs = np.array([(ptont[pp][ss]+1,ptont[pp][ss]-1) for pp in range(len(self.do_pts))])

    def dynamics_measures(self):
        #plot the *diff* between weeks
        rec_Zs = self.get_recZs()
        
        plt.figure()
        for ss,side in enumerate(['Left','Right']):
            plt.subplot(3,2,(ss+1)+0)
            plt.plot((self.Zdiff[side]))
            
            plt.subplot(3,2,(ss+1)+2)
            abs_Zdiff = np.abs(np.diff(self.Zdiff[side],axis=0))
            plt.plot(abs_Zdiff,alpha=0.2)
            plt.plot(np.nanmean(abs_Zdiff,axis=1),color='black')
        
        #plot the histogram of impedances, all of them
        plt.figure()
        for ss,side in enumerate(['Left','Right']):
            plt.subplot(1,2,ss+1)
            #stack everything together
            side_stack = np.vstack((rec_Zs[side][:,0,:],rec_Zs[side][:,1,:])).reshape(-1,1)
            side_stack = side_stack[~np.isnan(side_stack)]
            plt.hist(side_stack)
            plt.vlines(np.median(side_stack),0,100)
            print(side + ' ' + str(np.median(side_stack)))
            
    def gen_Zdiff(self):
        self.Zdiff = {'Left':[],'Right':[]}
        rec_Zs = self.get_recZs()
        for side in ['Left','Right']:
            self.Zdiff[side] = np.squeeze(rec_Zs[side][:,1,:].T - rec_Zs[side][:,0,:].T)
        

    def plot_recZs(self):
        rec_Zs = self.get_recZs()
        plt.figure()
        for ss,side in enumerate(['Left','Right']):
            plt.subplot(3,2,(ss+1)+0)
            plt.plot(rec_Zs[side][:,0,:].T)
            plt.ylim((500,2000))
            
            plt.subplot(3,2,(ss+1)+2)
            plt.plot(rec_Zs[side][:,1,:].T)
            plt.ylim((500,2000))
            
            plt.subplot(3,2,(ss+1)+4)
            plt.plot(rec_Zs[side][:,1,:].T - rec_Zs[side][:,0,:].T)
            plt.legend(self.pts)
            plt.ylim((-600,600))
        

#%%

if __name__ == '__main__':
    print('Doing Zs')
    Z_lib = Z_class()         
    Z_lib.get_recZs()
    #%%
    Z_lib.plot_recZs()
    Z_lib.dynamics_measures()