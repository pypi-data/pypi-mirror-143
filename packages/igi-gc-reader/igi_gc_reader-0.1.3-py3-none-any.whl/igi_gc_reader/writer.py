from typing import List

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.workbook.workbook import Workbook

from igi_gc_reader.reader.gc_reader import IGcSheet, GcSheet, GcSheetError
from igi_gc_reader.sample_data import IGIUserFriendlyException, SampleData
from igi_gc_reader.reader.excel_address import col_index_to_letter
from igi_gc_reader.utils import gc_logger


def write_sheets(gc_sheets: List[IGcSheet], out_path: str):

    def write_sheet(wb: Workbook, idx: int, gc: GcSheet):
        sh = wb.create_sheet(index=idx, title=gc.sheet_name)
        sample_data = SampleData(gc.page_data, igi_analysis=gc.get_igi_analysis())

        context_dict = gc.context_data.get_data_dict()
        value_row = sample_data.n_header_rows + 1
        for col_idx, (header, value) in enumerate(context_dict.items()):
            col = col_index_to_letter(col_idx)
            sh[f"{col}1"] = header
            sh[f"{col}{value_row}"] = value

        headers = list(sample_data.get_headers())
        sample_row = list(sample_data.get_sample_row())
        for col_idx, ((header_1, header_2), value) in enumerate(zip(headers, sample_row), 
                                                                start=col_idx+1):
            col = col_index_to_letter(col_idx)
            sh[f"{col}1"] = header_1
            if header_2 is None:
                sh[f"{col}2"] = value
            else:
                sh[f"{col}2"] = header_2
                sh[f"{col}3"] = value

    def write_error(wb: Workbook, idx: int, gc: GcSheetError):
        sh = wb.create_sheet(index=idx, title=f"{gc.sheet_name}_err")
        sh["A1"] = "Error reading GC data:"
        sh["A1"].font = Font(bold=True)
        sh["A2"] = str(gc.error)
        sh["A2"].alignment = Alignment(wrap_text=True, vertical="top")
        sh.row_dimensions[2].height = 200
        sh.column_dimensions["A"].width = 125

        # rmv original sheet if it failed to write
        if gc.sheet_name in wb.sheetnames:
            del wb[gc.sheet_name]

    wb = openpyxl.Workbook()    
    # starting out with 1 sheet out per sheet in - can comb later if required
    existing_sht = wb.sheetnames[0]

    for idx, gc in enumerate(gc_sheets):
        try:
            gc_logger.info(f"  writing sheet data for: {gc.sheet_name}")
            if isinstance(gc, GcSheet):
                write_sheet(wb, idx, gc)
            elif isinstance(gc, GcSheetError):
                write_error(wb, idx, gc)
        except IGIUserFriendlyException as e:  # catch any errors that don't come up until writing
            gc_err = GcSheetError(gc.sheet_name, gc.file_class, e)
            write_error(wb, idx, gc_err)

    del wb[existing_sht]
    gc_logger.info(f"\nSaving output to: {out_path}")
    wb.save(out_path)
